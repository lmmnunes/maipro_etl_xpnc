
import pandas as pd
import datetime as dt
import os
import pandas as pd
from IPython.display import display
from openpyxl import load_workbook
import numpy as np
import params


def get_dict(agency, type, data):

    source_file = type
    cols = data.columns

    # stats base
    dic = data.describe(percentiles=[.02, .05, .25, .5, .75, .95, .98]).reset_index()[1:]

    print(dt.datetime.now().strftime("%H:%M:%S"), "Base stats done")

    # Não nulos
    freq = pd.DataFrame([data[cols[0]].count()], columns=[cols[0]])
    for c in cols[1:]:
        freq = freq.merge(pd.DataFrame([data[c].count()], columns=[c]), how="outer", left_index=True, right_index=True)
    freq.insert(0 ,"index" ,"# Não nulos")
    dic = dic.append(freq)

    print(dt.datetime.now().strftime("%H:%M:%S"), "# nulls done")

    # % Não nulos
    freq = pd.DataFrame([float('{:.2f}'.format((data[cols[0]].count() / len(data) ) *100))], columns=[cols[0]])
    for c in cols[1:]:
        freq = freq.merge(pd.DataFrame([float('{:.2f}'.format((data[c].count() / len(data) ) *100))], columns=[c]), how="outer", left_index=True, right_index=True)
    freq.insert(0 ,"index" ,"% Não nulos")
    dic = dic.append(freq)

    print(dt.datetime.now().strftime("%H:%M:%S"), "% nulls done")

    # outras stats
    more_stats = {"# Nulos": data.isnull().sum(),
                  "% Nulos": data.isnull().mean().apply(lambda n: float('{:.2f}'.format( n *100))),
                  "Valores distintos": data.nunique(),
                  "Tipos": data.dtypes}

    for k, v in more_stats.items():
        dic = dic.append(pd.DataFrame(v, columns=[k]).T.reset_index())

    print(dt.datetime.now().strftime("%H:%M:%S"), "Other stats done")

    # Frequência top 20 desc
    freq = pd.DataFrame([str(data[cols[0]].value_counts().sort_values(ascending=False).to_dict())], columns=[cols[0]])
    for c in cols[1:]:
        freq = freq.merge \
            (pd.DataFrame([str(data[c].value_counts().sort_values(ascending=False).to_dict())], columns=[c]), how="outer", left_index=True, right_index=True)
    freq.insert(0 ,"index" ,"Mais frequentes top 20")
    dic = dic.append(freq)

    print(dt.datetime.now().strftime("%H:%M:%S"), "Freq desc done")

    # Frequência top 20 asc
    freq = pd.DataFrame([str(data[cols[0]].value_counts().sort_values(ascending=True).to_dict())], columns=[cols[0]])
    for c in cols[1:]:
        freq = freq.merge \
            (pd.DataFrame([str(data[c].value_counts().sort_values(ascending=True).to_dict())], columns=[c]), how="outer", left_index=True, right_index=True)
    freq.insert(0 ,"index" ,"Menos frequentes top 20")
    dic = dic.append(freq)

    print(dt.datetime.now().strftime("%H:%M:%S"), "Freq asc done")

    # Mais frequente - desc
    freq = pd.DataFrame([str(data[cols[0]].value_counts().sort_values(ascending=False).head(1).to_dict())], columns=[cols[0]])
    for c in cols[1:]:
        freq = freq.merge \
            (pd.DataFrame([str(data[c].value_counts().sort_values(ascending=False).head(1).to_dict())], columns=[c]), how="outer", left_index=True, right_index=True)
    freq.insert(0 ,"index" ,"# Mais frequente")
    dic = dic.append(freq)

    print(dt.datetime.now().strftime("%H:%M:%S"), "Mais freq done")

    # Menos frequente - asc
    freq = pd.DataFrame([str(data[cols[0]].value_counts().sort_values(ascending=True).head(1).to_dict())], columns=[cols[0]])
    for c in cols[1:]:
        freq = freq.merge \
            (pd.DataFrame([str(data[c].value_counts().sort_values(ascending=True).head(1).to_dict())], columns=[c]), how="outer", left_index=True, right_index=True)
    freq.insert(0 ,"index" ,"# Menos frequente")
    dic = dic.append(freq)

    print(dt.datetime.now().strftime("%H:%M:%S"), "Menos freq done")


    # % mais e menos freq
    n_rows = data.shape[0]

    def func(n):
        display(n)
        # n = eval(n.iloc[0])
        # n = list(n.values())[0]
        n = str(n.iloc[0]).split(":")[-1].replace("}" ,"").replace("{" ,"").strip()
        display(n)
        if n == '':
            n = 100
        else:
            n = (float(n) / n_rows) * 100
        n = float('{:.2f}'.format(n))
        return n

    display(dic)

    temp1 = dic.iloc[[-2] ,1:].apply(lambda n: func(n))
    temp2 = dic.iloc[[-1] ,1:].apply(lambda n: func(n))
    temp1 = pd.Series(["% Mais frequente"]).append(temp1)
    temp2 = pd.Series(["% Menos frequente"]).append(temp2)

    dic = dic.append(pd.DataFrame([temp1], columns=cols))
    dic = dic.append(pd.DataFrame([temp2], columns=cols))
    dic.iloc[[-2] ,0] = "% Mais frequente"
    dic.iloc[[-1] ,0] = "% Menos frequente"

    display(dic)

    print(dt.datetime.now().strftime("%H:%M:%S"), "% Mais e menos freq done")

    # inverter
    dic_t = dic.reset_index(drop=True).T

    # arranjar cols
    names = dic_t.iloc[0]
    dict_names_temp = dict(zip(dic_t.columns, names))
    dic_renamed_columns = dic_t.rename(columns=dict_names_temp)
    dic_renamed_columns.columns.names = ['Nº']
    dic_renamed_columns = dic_renamed_columns.reset_index()
    dict_names_final = {"index" :"Campo" ,"mean" :"Média", "std" :"Desvio-padrão", "min" :"Mínimo", "max" :"Máximo", "50%" :"Mediana"}
    dic_renamed_columns_final = dic_renamed_columns.iloc[1:].rename(columns=dict_names_final)

    dic_f = dic_renamed_columns_final

    dic_f.insert(0 ,"Fonte" ,source_file)

    print(dt.datetime.now().strftime("%H:%M:%S"), "Invert and fixes done")

    # tirar info da documentacao
    map_etl_url = params.metadata
    map_etl = pd.read_excel(map_etl_url ,engine='openpyxl')
    map_etl_1 = map_etl[["Ficheiro Fonte" ,"Campo Fonte" ,"Tipo Fonte" ,"Descrição Fonte" ,"Agência" ,"Nível de Análise"]]
    map_etl_2 = map_etl[["Ficheiro Destino" ,"Campo Destino" ,"Tipo Destino" ,"Descrição Destino" ,"Agência" ,"Nível de Análise"
         ,"Pode ser Feature"]]
    map_etl_1.rename \
        (columns=dict(zip(map_etl_1.columns ,[c.split(' ')[0] for c in map_etl_1.columns[0:5] ] +["Nível de Análise"])), inplace=True)
    map_etl_2.rename(columns=dict(zip(map_etl_2.columns
                                      ,[c.split(' ')[0] for c in map_etl_2.columns[0:5] ] +["Nível de Análise"
                                                                                           ,"Pode ser Feature"])), inplace=True)
    map_etl_appended = map_etl_1.append(map_etl_2)
    map_etl_cleaned = map_etl_appended.drop_duplicates()
    map_etl_selected = map_etl_cleaned[((map_etl_cleaned['Ficheiro' ]==source_file) &
                ((map_etl_cleaned['Agência'] == "Ambas") | (map_etl_cleaned['Agência'] == agency.upper())))]
    map_etl_selected["Pode ser Feature"] = map_etl_selected["Pode ser Feature"].fillna(0)

    dic_f.insert(2, "Descrição", dic_f["Campo"].map(dict(zip(map_etl_cleaned["Campo"], map_etl_cleaned["Descrição"]))))
    dic_f.insert(3, "Tipo", dic_f["Campo"].map(dict(zip(map_etl_cleaned["Campo"], map_etl_cleaned["Tipo"]))))
    dic_f.insert(4, "Pode ser Feature",
                 dic_f["Campo"].map(dict(zip(map_etl_cleaned["Campo"], map_etl_cleaned["Pode ser Feature"]))))
    dic_f.insert(5, "Agência", dic_f["Campo"].map(dict(zip(map_etl_cleaned["Campo"], map_etl_cleaned["Agência"]))))
    dic_f.insert(6, "Nível de Análise",
                 dic_f["Campo"].map(dict(zip(map_etl_cleaned["Campo"], map_etl_cleaned["Nível de Análise"]))))

    dic_f.columns.names = ['Número']

    print(dt.datetime.now().strftime("%H:%M:%S"), "All stats done")

    return dic_f


def append_df_to_excel(filename, df, sheet_name='Sheet1', startrow=None,
                       truncate_sheet=False,
                       **to_excel_kwargs):
    # Excel file doesn't exist - saving and exiting
    if not os.path.isfile(filename):
        df.to_excel(
            filename,
            sheet_name=sheet_name,
            startrow=startrow if startrow is not None else 0,
            **to_excel_kwargs)
        return

    # ignore [engine] parameter if it was passed
    if 'engine' in to_excel_kwargs:
        to_excel_kwargs.pop('engine')

    writer = pd.ExcelWriter(filename, engine='openpyxl', mode='a')

    # try to open an existing workbook
    writer.book = load_workbook(filename)

    # get the last row in the existing Excel sheet
    # if it was not specified explicitly
    if startrow is None and sheet_name in writer.book.sheetnames:
        startrow = writer.book[sheet_name].max_row

    # truncate sheet
    if truncate_sheet and sheet_name in writer.book.sheetnames:
        # index of [sheet_name] sheet
        idx = writer.book.sheetnames.index(sheet_name)
        # remove [sheet_name]
        writer.book.remove(writer.book.worksheets[idx])
        # create an empty sheet [sheet_name] using old index
        writer.book.create_sheet(sheet_name, idx)

    # copy existing sheets
    writer.sheets = {ws.title: ws for ws in writer.book.worksheets}

    if startrow is None:
        startrow = 0

    # write out the new sheet
    df.to_excel(writer, sheet_name, startrow=startrow, **to_excel_kwargs)

    # save the workbook
    writer.save()


def create_data_dictionary_folder(agency, type, dict_directory):
    if agency == params.iapmei:
        directory = params.iapmei_sources
        extra = ""
    else:
        directory = params.aicep_sources
        extra = "\\csv_dados\\"

    folders = os.listdir(directory)
    # comentar ou descomentar para selecionar pastas especificas
    # folders = ["candidaturas_dados"]

    for folder in folders:
        i = 0
        curr_directory = directory + "\\" + folder + extra + "\\"
        filenames = os.listdir(curr_directory)
        for filename in filenames:
            if ".csv" in filename:
                i = i + 1
                print(dt.datetime.now().strftime("%H:%M:%S"), "Exploring file", filename)
                data = pd.read_csv(curr_directory + filename)
                append_df_to_excel(r'' + dict_directory + 'dicionario.xlsx', get_dict(agency, type, data),
                                   sheet_name=agency + "_" + folder.split("_")[0], index=False, header=i == 1, encoding='utf-8-sig')
                print(dt.datetime.now().strftime("%H:%M:%S"), "Writing to excel done")


def create_data_dictionary_file(agency, type, path, dict_directory):
    name = path.split("\\")[-1].split(".")[0]
    print(dt.datetime.now().strftime("%H:%M:%S"), "Exploring file", name)

    # ler ficheiro
    try:
        data = pd.read_excel(path, engine='openpyxl')
    except:
        data = pd.read_csv(path)

    #criar dict e meter num excel
    append_df_to_excel(r'' + dict_directory + 'dicionario.xlsx', get_dict(agency, type, data),
                       sheet_name=name, index=False, encoding='utf-8-sig')
    print(dt.datetime.now().strftime("%H:%M:%S"), "Writing to excel done")


def create_data_dictionary(agency, type, path, create_dict_sources, create_dict_others):
    if create_dict_others or create_dict_sources:
        print(dt.datetime.now().strftime("%H:%M:%S"), "Starting data dictionary")
        dict_directory = params.save_root_folder
        if create_dict_others:
            create_data_dictionary_file(agency, type, path, dict_directory)
        if create_dict_sources:
            create_data_dictionary_folder(agency, type, dict_directory)

